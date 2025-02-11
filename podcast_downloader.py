# Description: This script downloads podcast episodes from an Excel file containing episode information.
import os
import requests
import openpyxl

def download_from_excel(excel_path='podcast_entries.xlsx', download_dir='podcast_episodes'):
    os.makedirs(download_dir, exist_ok=True)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    try:
        published_index = header.index("published")
        audio_url_index = header.index("audio_url")
    except ValueError as e:
        print(f"Header missing: {e}")
        return
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        published = row[published_index]
        audio_url = row[audio_url_index]
        if audio_url and published != "N/A":
            file_path = os.path.join(download_dir, f"{published}.mp3")
            try:
                response = requests.get(audio_url)
                response.raise_for_status()
                with open(file_path, 'wb') as out_file:
                    out_file.write(response.content)
                print(f"Downloaded: {file_path}")
            except Exception as e:
                print(f"Failed to download {audio_url}: {e}")

if __name__ == '__main__':
    download_from_excel()
